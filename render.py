# 渲染Excel模板，输出的文件为多个文件或一个文件多个Sheet
# 模板文件 template.xlsx
# 数据文件 data.xlsx，第一行说明，第二行是目标单元格，第三行开始是要渲染的数据，每行对应一个输出文件（或一个Sheet）
# 输出路径 output，文件名在数据文件的第一列
import os
from typing import Dict
import openpyxl as xl

# 输出模式
# 多文件模式 1：数据文件中每一行对应一个输出文件
# 单文件模式 2：数据文件中每一行对应一个Sheet
MODE = 2

# 多文件输出模式下输出目录
_OUTPUT_DIR = "output/"

# 多Sheet模式下的输出文件名
_OUTPUT_FILE = "output/result.xlsx"

# 模板的Sheet名
_TPL_SHEET_NAME = '模板'


def build_target_map(data_wb) -> Dict:
    data_sheet = data_wb.active
    target_map = dict()
    for row in data_sheet.iter_rows(min_row=2, min_col=2, max_row=2, values_only=True):
        idx = 1
        for cell in row:
            if cell is None:
                break
            target_map[idx] = cell
            idx += 1
    return target_map


def render_and_save_separate_file(tpl_wb, data_row, target_map: Dict) -> None:
    file_name = _OUTPUT_DIR + data_row[0] + '.xlsx'
    print('渲染 {}'.format(file_name))
    sheet = tpl_wb.active
    sheet.title = 'Sheet1'
    render(sheet, data_row, target_map)
    tpl_wb.save(file_name)
    print('写入完成 {}'.format(file_name))


def render_to_new_sheet(tpl_wb, data_row, target_map: Dict) -> None:
    sheet_name = data_row[0]
    print('渲染Sheet {}'.format(sheet_name))
    sheet = tpl_wb.copy_worksheet(tpl_wb[_TPL_SHEET_NAME])
    sheet.title = sheet_name
    render(sheet, data_row, target_map)
    pass


def render(sheet, data_row, target_map) -> None:
    for idx, target in target_map.items():
        value = data_row[idx]
        sheet[target] = value


if __name__ == '__main__':
    # 读取模板
    print('读取template.xlsx...')
    tpl_wb = xl.load_workbook('template.xlsx')

    # 读取数据文件
    print('读取data.xlsx...')
    data_wb = xl.load_workbook('data.xlsx', data_only=True)

    # 解析填空位置
    target_map = build_target_map(data_wb)

    # 创建输出目录
    try:
        os.mkdir('output')
        print('创建输出目录 {}'.format(_OUTPUT_DIR))
    except:
        print('输出目录已存在')

    # 执行渲染
    for row in data_wb.active.iter_rows(min_row=3, min_col=1, values_only=True):
        if MODE == 1:
            render_and_save_separate_file(tpl_wb, row, target_map)
        elif MODE == 2:
            render_to_new_sheet(tpl_wb, row, target_map)

    # 如果是模式2，那么最后要删除模板Sheet，并且保存
    if MODE == 2:
        del tpl_wb[_TPL_SHEET_NAME]
        tpl_wb.save(_OUTPUT_FILE)
        print('写入完成 {}'.format(_OUTPUT_FILE))

    print('全部完成')
